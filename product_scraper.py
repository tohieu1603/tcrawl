#!/usr/bin/env python3
"""
Product Web Scraper
Cào dữ liệu sản phẩm từ các trang web e-commerce và lưu vào Excel

Hỗ trợ:
- Dienmayxanh (DMX)
- Thegioididong (TGDD)
- Cellphones
- FPT Shop
- Generic sites

Usage:
    python product_scraper.py <url> [--output <file.xlsx>]
    python product_scraper.py --help
"""

import argparse
import json
import os
import re
import sys
import time
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional
from urllib.parse import urlparse

try:
    import requests
    from bs4 import BeautifulSoup
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError as e:
    print(f"Error: Missing required package. Please install dependencies:")
    print("pip install requests beautifulsoup4 openpyxl lxml")
    sys.exit(1)

# Optional Selenium support for JS-rendered pages
SELENIUM_AVAILABLE = False
try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    SELENIUM_AVAILABLE = True
except ImportError:
    pass


@dataclass
class ProductData:
    """Dữ liệu sản phẩm được cào"""
    name: str
    sku_prefix: str = ""
    slug: str = ""
    brand_name: str = ""
    category_name: str = ""
    base_price: float = 0
    compare_at_price: float = 0
    short_description: str = ""
    description: str = ""
    is_featured: bool = False
    status: str = "draft"
    meta_title: str = ""
    meta_description: str = ""
    tags: str = ""

    # Variants
    variants: List[Dict[str, Any]] = field(default_factory=list)

    # Attributes/Specs
    attributes: List[Dict[str, Any]] = field(default_factory=list)

    # Media
    images: List[str] = field(default_factory=list)

    # Source info
    source_url: str = ""
    scraped_at: str = ""


class BaseScraper(ABC):
    """Base class cho các scraper"""

    def __init__(self, use_selenium: bool = False):
        self.use_selenium = use_selenium and SELENIUM_AVAILABLE
        self.driver = None
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
            'Accept-Language': 'vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-User': '?1',
            'Cache-Control': 'max-age=0',
        })

    def _init_selenium(self):
        """Initialize Selenium WebDriver"""
        if not SELENIUM_AVAILABLE:
            print("Selenium not available. Install: pip install selenium")
            return None
        if self.driver is None:
            options = Options()
            options.add_argument('--headless')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--disable-gpu')
            options.add_argument('--window-size=1920,1080')
            options.add_argument('--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36')
            self.driver = webdriver.Chrome(options=options)
        return self.driver

    def _close_selenium(self):
        """Close Selenium WebDriver"""
        if self.driver:
            self.driver.quit()
            self.driver = None

    @abstractmethod
    def can_handle(self, url: str) -> bool:
        """Kiểm tra scraper có hỗ trợ URL này không"""
        pass

    @abstractmethod
    def scrape(self, url: str) -> ProductData:
        """Cào dữ liệu sản phẩm từ URL"""
        pass

    def fetch_page(self, url: str, retries: int = 3) -> BeautifulSoup:
        """Fetch và parse HTML page với retry"""
        # Use Selenium if enabled (for JS-rendered pages)
        if self.use_selenium:
            return self._fetch_with_selenium(url)

        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

        for attempt in range(retries):
            try:
                # Add referer header based on domain
                from urllib.parse import urlparse
                parsed = urlparse(url)
                self.session.headers['Referer'] = f"{parsed.scheme}://{parsed.netloc}/"

                response = self.session.get(url, timeout=30, verify=False)
                response.raise_for_status()
                return BeautifulSoup(response.text, 'lxml')
            except Exception as e:
                if attempt < retries - 1:
                    print(f"Retry {attempt + 1}/{retries} after error: {e}")
                    import time
                    time.sleep(2)
                else:
                    print(f"Error fetching {url}: {e}")
                    raise

    def _fetch_with_selenium(self, url: str) -> BeautifulSoup:
        """Fetch page using Selenium for JS-rendered content"""
        driver = self._init_selenium()
        if not driver:
            raise RuntimeError("Selenium driver not available")

        try:
            print(f"  Using Selenium to fetch {url}...")
            driver.get(url)
            # Wait for gallery images to load
            time.sleep(3)
            # Try to wait for owl-dots to appear
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".owl-dots, .gallery, .slider-product"))
                )
            except Exception:
                pass  # Continue even if not found
            html = driver.page_source
            return BeautifulSoup(html, 'lxml')
        except Exception as e:
            print(f"Selenium error: {e}")
            raise

    def clean_text(self, text: Optional[str]) -> str:
        """Clean và normalize text"""
        if not text:
            return ""
        return re.sub(r'\s+', ' ', text.strip())

    def clean_price(self, price_text: Optional[str]) -> float:
        """Extract số từ chuỗi giá"""
        if not price_text:
            return 0
        # Remove tất cả ký tự không phải số
        numbers = re.findall(r'\d+', price_text.replace('.', '').replace(',', ''))
        if numbers:
            return float(''.join(numbers))
        return 0

    def generate_sku(self, name: str) -> str:
        """Generate SKU prefix từ tên sản phẩm"""
        # Lấy chữ cái đầu của mỗi từ + timestamp
        words = name.split()[:3]
        prefix = ''.join(w[0].upper() for w in words if w)
        timestamp = datetime.now().strftime('%y%m%d%H%M')
        return f"{prefix}-{timestamp}"

    def generate_slug(self, name: str) -> str:
        """Generate slug từ tên sản phẩm"""
        # Convert to lowercase, remove Vietnamese diacritics
        slug = name.lower()
        vietnamese_map = {
            'à': 'a', 'á': 'a', 'ả': 'a', 'ã': 'a', 'ạ': 'a',
            'ă': 'a', 'ằ': 'a', 'ắ': 'a', 'ẳ': 'a', 'ẵ': 'a', 'ặ': 'a',
            'â': 'a', 'ầ': 'a', 'ấ': 'a', 'ẩ': 'a', 'ẫ': 'a', 'ậ': 'a',
            'đ': 'd',
            'è': 'e', 'é': 'e', 'ẻ': 'e', 'ẽ': 'e', 'ẹ': 'e',
            'ê': 'e', 'ề': 'e', 'ế': 'e', 'ể': 'e', 'ễ': 'e', 'ệ': 'e',
            'ì': 'i', 'í': 'i', 'ỉ': 'i', 'ĩ': 'i', 'ị': 'i',
            'ò': 'o', 'ó': 'o', 'ỏ': 'o', 'õ': 'o', 'ọ': 'o',
            'ô': 'o', 'ồ': 'o', 'ố': 'o', 'ổ': 'o', 'ỗ': 'o', 'ộ': 'o',
            'ơ': 'o', 'ờ': 'o', 'ớ': 'o', 'ở': 'o', 'ỡ': 'o', 'ợ': 'o',
            'ù': 'u', 'ú': 'u', 'ủ': 'u', 'ũ': 'u', 'ụ': 'u',
            'ư': 'u', 'ừ': 'u', 'ứ': 'u', 'ử': 'u', 'ữ': 'u', 'ự': 'u',
            'ỳ': 'y', 'ý': 'y', 'ỷ': 'y', 'ỹ': 'y', 'ỵ': 'y',
        }
        for vn, en in vietnamese_map.items():
            slug = slug.replace(vn, en)
        # Replace non-alphanumeric with hyphen
        slug = re.sub(r'[^a-z0-9]+', '-', slug)
        # Remove leading/trailing hyphens
        slug = slug.strip('-')
        return slug


class DienmayxanhScraper(BaseScraper):
    """Scraper cho Dienmayxanh.com và Thegioididong.com"""

    def can_handle(self, url: str) -> bool:
        domain = urlparse(url).netloc.lower()
        return any(d in domain for d in ['dienmayxanh.com', 'thegioididong.com'])

    def scrape(self, url: str) -> ProductData:
        soup = self.fetch_page(url)
        html_text = str(soup)
        product = ProductData(
            name="",
            source_url=url,
            scraped_at=datetime.now().isoformat()
        )

        # Tên sản phẩm
        name_elem = soup.select_one('h1')
        if name_elem:
            product.name = self.clean_text(name_elem.text)
            product.slug = self.generate_slug(product.name)
            product.sku_prefix = self.generate_sku(product.name)

        # Giá - thử nhiều cách
        # 1. Từ JSON embedded
        price_match = re.search(r'"price"\s*:\s*(\d+)', html_text)
        if price_match:
            product.base_price = float(price_match.group(1))

        # 2. Từ data-price attribute (giá lớn nhất hợp lý - thường là giá SP chính)
        if product.base_price == 0:
            data_prices = re.findall(r'data-price="(\d+)"', html_text)
            valid_prices = [float(p) for p in data_prices if float(p) > 1000000]  # > 1 triệu
            if valid_prices:
                product.base_price = max(valid_prices)

        # 3. Từ HTML selectors
        if product.base_price == 0:
            price_selectors = [
                '.box-price .box-price-present', '.product-price .present',
                '.bs_price', '.price', '[class*="price-current"]'
            ]
            for selector in price_selectors:
                price_elem = soup.select_one(selector)
                if price_elem:
                    price = self.clean_price(price_elem.text)
                    if price > 1000000:  # > 1 triệu
                        product.base_price = price
                        break

        # Giá gốc (nếu có)
        old_price_selectors = ['.box-price .box-price-old', '.product-price .old', '.price-old']
        for selector in old_price_selectors:
            old_price_elem = soup.select_one(selector)
            if old_price_elem:
                product.compare_at_price = self.clean_price(old_price_elem.text)
                break

        # Brand từ breadcrumb hoặc title
        brand_elem = soup.select_one('.breadcrumb a:nth-child(2), .box04.box-brand a')
        if brand_elem:
            product.brand_name = self.clean_text(brand_elem.text)
        # Fallback: extract từ tên sản phẩm
        if not product.brand_name and product.name:
            known_brands = ['iPhone', 'Samsung', 'Xiaomi', 'OPPO', 'Vivo', 'Realme', 'Sony', 'LG', 'Apple']
            for brand in known_brands:
                if brand.lower() in product.name.lower():
                    product.brand_name = brand
                    break

        # Category
        cat_elem = soup.select_one('.breadcrumb a:nth-child(1)')
        if cat_elem:
            product.category_name = self.clean_text(cat_elem.text)

        # Short description - từ highlights
        highlight_items = soup.select('.box-content ul li, .highlight li, .box-specifi li')
        if highlight_items:
            highlights = [self.clean_text(li.text) for li in highlight_items[:5]]
            product.short_description = ' | '.join(highlights)[:500]

        # Description
        desc_elem = soup.select_one('.box-des article, .article-content, .product-article')
        if desc_elem:
            product.description = str(desc_elem)

        # Images - CHỈ lấy từ gallery sản phẩm (owl-dots)
        # Priority 1: owl-dots thumbnails (gallery chính)
        owl_dots = soup.select('.owl-dots button img, .owl-dots .owl-dot img')
        if owl_dots:
            for img in owl_dots:
                src = img.get('src') or img.get('data-src')
                if src:
                    # Convert thumbnail URL to full size
                    # Pattern: https://img.tgdd.vn/imgt/f_webp.../https://cdn.tgdd.vn/...180x120.jpg
                    # -> https://cdn.tgdd.vn/...jpg (original)
                    if 'imgt/' in src and 'https://cdn' in src:
                        # Extract original URL from wrapper
                        match = re.search(r'(https://cdn[^,\s]+)', src)
                        if match:
                            src = match.group(1)
                    # Remove SMALL thumbnail size suffix only (180x120, 100x100, etc.)
                    # Keep large sizes like 1920x1080 (actual image)
                    size_match = re.search(r'-?(\d+)x(\d+)(?=\.\w+$)', src)
                    if size_match:
                        w, h = int(size_match.group(1)), int(size_match.group(2))
                        # Only remove if it's a small thumbnail (< 500px)
                        if w < 500 and h < 500:
                            src = re.sub(r'-?\d+x\d+(?=\.\w+$)', '', src)
                    # Handle cdnv2 URLs
                    src = src.replace('cdnv2.tgdd.vn/mwg-static/dmx/', 'cdn.tgdd.vn/')
                    if src not in product.images and 'icon' not in src.lower():
                        product.images.append(src)

        # Priority 2: Main slider image (ảnh lớn đầu tiên)
        if not product.images:
            main_img = soup.select_one('.owl-carousel .owl-item.active img, .slide-product img')
            if main_img:
                src = main_img.get('src') or main_img.get('data-src')
                if src and src not in product.images:
                    product.images.append(src)

        # Priority 3: Fallback - ảnh từ gallery container cụ thể
        if not product.images:
            gallery_selectors = ['.gallery-product img', '.slider-product img', '.product-slide img']
            for selector in gallery_selectors:
                img_elems = soup.select(selector)
                for img in img_elems[:10]:  # Limit 10 ảnh
                    src = img.get('data-src') or img.get('src')
                    if src and ('cdn' in src or 'http' in src):
                        # Only remove small thumbnail sizes
                        size_match = re.search(r'-?(\d+)x(\d+)(?=\.\w+$)', src)
                        if size_match:
                            w, h = int(size_match.group(1)), int(size_match.group(2))
                            if w < 500 and h < 500:
                                src = re.sub(r'-?\d+x\d+(?=\.\w+$)', '', src)
                        if src not in product.images and 'icon' not in src.lower():
                            product.images.append(src)
                if product.images:
                    break

        # Thông số kỹ thuật - nhiều patterns
        spec_selectors = [
            '.parameter li', '.box-specifi li', '.specifications tr',
            '.technical li', '.spec-list li', 'ul.specifi li'
        ]
        display_order = 1
        for selector in spec_selectors:
            spec_items = soup.select(selector)
            for item in spec_items:
                # Pattern 1: separate elements
                name_elem = item.select_one('.tit, .name, .label, td:first-child, span:first-child')
                value_elem = item.select_one('.result, .value, td:last-child, span:last-child')

                if name_elem and value_elem and name_elem != value_elem:
                    attr_name = self.clean_text(name_elem.text).rstrip(':')
                    attr_value = self.clean_text(value_elem.text)
                    if attr_name and attr_value and attr_name != attr_value:
                        product.attributes.append({
                            'attribute_name': attr_name,
                            'value': attr_value,
                            'display_group': 'Thông số kỹ thuật',
                            'display_order': display_order
                        })
                        display_order += 1
                else:
                    # Pattern 2: text with colon separator
                    text = self.clean_text(item.text)
                    if ':' in text:
                        parts = text.split(':', 1)
                        if len(parts) == 2:
                            attr_name = parts[0].strip()
                            attr_value = parts[1].strip()
                            if attr_name and attr_value:
                                product.attributes.append({
                                    'attribute_name': attr_name,
                                    'value': attr_value,
                                    'display_group': 'Thông số kỹ thuật',
                                    'display_order': display_order
                                })
                                display_order += 1

        # Variants (màu sắc, dung lượng)
        variant_selectors = ['.box-color a', '.list-color a', '.box-choose a', '.choose-attr a']
        for selector in variant_selectors:
            variant_elems = soup.select(selector)
            for i, var_elem in enumerate(variant_elems):
                var_name = self.clean_text(var_elem.get('title') or var_elem.text)
                var_price = self.clean_price(var_elem.get('data-price') or '')
                if var_name and var_name not in ['', ' ']:
                    product.variants.append({
                        'sku': f"{product.sku_prefix}-V{i+1}",
                        'name': f"{product.name} - {var_name}",
                        'price': var_price or product.base_price,
                        'option_1_type': 'Phiên bản',
                        'option_1_value': var_name,
                        'is_default': i == 0
                    })

        # Nếu không có variants, tạo 1 variant mặc định
        if not product.variants:
            product.variants.append({
                'sku': f"{product.sku_prefix}-01",
                'name': product.name,
                'price': product.base_price,
                'is_default': True
            })

        return product


class CellphonesScraper(BaseScraper):
    """Scraper cho Cellphones.com.vn"""

    def can_handle(self, url: str) -> bool:
        return 'cellphones.com.vn' in urlparse(url).netloc.lower()

    def scrape(self, url: str) -> ProductData:
        soup = self.fetch_page(url)
        product = ProductData(
            name="",
            source_url=url,
            scraped_at=datetime.now().isoformat()
        )

        # Tên sản phẩm
        name_elem = soup.select_one('h1')
        if name_elem:
            product.name = self.clean_text(name_elem.text)
            product.slug = self.generate_slug(product.name)
            product.sku_prefix = self.generate_sku(product.name)

        # Giá
        price_elem = soup.select_one('.product__price--show, .tpt---sale-price')
        if price_elem:
            product.base_price = self.clean_price(price_elem.text)

        # Giá gốc
        old_price_elem = soup.select_one('.product__price--through, .tpt---list-price')
        if old_price_elem:
            product.compare_at_price = self.clean_price(old_price_elem.text)

        # Brand
        brand_elem = soup.select_one('.breadcrumb-item:nth-child(2) a')
        if brand_elem:
            product.brand_name = self.clean_text(brand_elem.text)

        # Description
        desc_elem = soup.select_one('.block-content-article, .product-detail')
        if desc_elem:
            product.description = str(desc_elem)

        # Images
        img_elems = soup.select('.gallery-product img, .swiper-slide img')
        for img in img_elems:
            src = img.get('data-src') or img.get('src')
            if src and 'http' in src and src not in product.images:
                product.images.append(src)

        # Thông số kỹ thuật
        spec_items = soup.select('.technical-content li, .specifications-item')
        display_order = 1
        for item in spec_items:
            name_elem = item.select_one('.title, span:first-child')
            value_elem = item.select_one('.value, span:last-child')
            if name_elem and value_elem:
                attr_name = self.clean_text(name_elem.text).rstrip(':')
                attr_value = self.clean_text(value_elem.text)
                if attr_name and attr_value:
                    product.attributes.append({
                        'attribute_name': attr_name,
                        'value': attr_value,
                        'display_group': 'Thông số kỹ thuật',
                        'display_order': display_order
                    })
                    display_order += 1

        # Default variant
        product.variants.append({
            'sku': f"{product.sku_prefix}-01",
            'name': product.name,
            'price': product.base_price,
            'is_default': True
        })

        return product


class FPTShopScraper(BaseScraper):
    """Scraper cho FPTShop.com.vn"""

    def can_handle(self, url: str) -> bool:
        return 'fptshop.com.vn' in urlparse(url).netloc.lower()

    def scrape(self, url: str) -> ProductData:
        soup = self.fetch_page(url)
        product = ProductData(
            name="",
            source_url=url,
            scraped_at=datetime.now().isoformat()
        )

        # Tên sản phẩm
        name_elem = soup.select_one('h1.st-name, h1')
        if name_elem:
            product.name = self.clean_text(name_elem.text)
            product.slug = self.generate_slug(product.name)
            product.sku_prefix = self.generate_sku(product.name)

        # Giá
        price_elem = soup.select_one('.st-price-main, .price-value')
        if price_elem:
            product.base_price = self.clean_price(price_elem.text)

        # Giá gốc
        old_price_elem = soup.select_one('.st-price-sub, .price-old')
        if old_price_elem:
            product.compare_at_price = self.clean_price(old_price_elem.text)

        # Brand từ breadcrumb
        brand_elem = soup.select_one('.breadcrumb a:nth-child(2)')
        if brand_elem:
            product.brand_name = self.clean_text(brand_elem.text)

        # Images
        img_elems = soup.select('.owl-carousel img, .product-gallery img')
        for img in img_elems:
            src = img.get('data-src') or img.get('src')
            if src and 'http' in src and src not in product.images:
                product.images.append(src)

        # Thông số kỹ thuật
        spec_items = soup.select('.st-param tr, .specification tr')
        display_order = 1
        for item in spec_items:
            cells = item.select('td')
            if len(cells) >= 2:
                attr_name = self.clean_text(cells[0].text).rstrip(':')
                attr_value = self.clean_text(cells[1].text)
                if attr_name and attr_value:
                    product.attributes.append({
                        'attribute_name': attr_name,
                        'value': attr_value,
                        'display_group': 'Thông số kỹ thuật',
                        'display_order': display_order
                    })
                    display_order += 1

        # Default variant
        product.variants.append({
            'sku': f"{product.sku_prefix}-01",
            'name': product.name,
            'price': product.base_price,
            'is_default': True
        })

        return product


class GenericScraper(BaseScraper):
    """Generic scraper cho các trang web khác"""

    def can_handle(self, url: str) -> bool:
        return True  # Fallback scraper

    def scrape(self, url: str) -> ProductData:
        soup = self.fetch_page(url)
        product = ProductData(
            name="",
            source_url=url,
            scraped_at=datetime.now().isoformat()
        )

        # Try common patterns for product name
        name_selectors = [
            'h1.product-title', 'h1.product-name', 'h1.title',
            '.product-title h1', '.product-name h1', 'h1'
        ]
        for selector in name_selectors:
            elem = soup.select_one(selector)
            if elem and elem.text.strip():
                product.name = self.clean_text(elem.text)
                break

        if product.name:
            product.slug = self.generate_slug(product.name)
            product.sku_prefix = self.generate_sku(product.name)

        # Try common patterns for price
        price_selectors = [
            '.product-price .current', '.price-current', '.sale-price',
            '.product-price', '.price', '[itemprop="price"]'
        ]
        for selector in price_selectors:
            elem = soup.select_one(selector)
            if elem:
                price = self.clean_price(elem.text or elem.get('content', ''))
                if price > 0:
                    product.base_price = price
                    break

        # Try to find images
        img_selectors = [
            '.product-gallery img', '.product-images img',
            '.gallery img', '.slider img', '[itemprop="image"]'
        ]
        for selector in img_selectors:
            imgs = soup.select(selector)
            for img in imgs:
                src = img.get('data-src') or img.get('src') or img.get('content')
                if src and 'http' in src and src not in product.images:
                    product.images.append(src)

        # Try to find description
        desc_selectors = [
            '.product-description', '.description', '[itemprop="description"]',
            '.product-detail', '.product-content'
        ]
        for selector in desc_selectors:
            elem = soup.select_one(selector)
            if elem:
                product.description = str(elem)
                break

        # Default variant
        if product.name:
            product.variants.append({
                'sku': f"{product.sku_prefix}-01",
                'name': product.name,
                'price': product.base_price,
                'is_default': True
            })

        return product


class ProductScraperManager:
    """Manager để chọn scraper phù hợp"""

    def __init__(self, use_selenium: bool = False):
        self.use_selenium = use_selenium
        self.scrapers: List[BaseScraper] = [
            DienmayxanhScraper(use_selenium=use_selenium),
            CellphonesScraper(use_selenium=use_selenium),
            FPTShopScraper(use_selenium=use_selenium),
            GenericScraper(use_selenium=use_selenium),  # Fallback
        ]

    def scrape(self, url: str) -> ProductData:
        """Cào dữ liệu từ URL"""
        for scraper in self.scrapers:
            if scraper.can_handle(url):
                print(f"Using scraper: {scraper.__class__.__name__}")
                if self.use_selenium:
                    print("  (with Selenium for JS content)")
                result = scraper.scrape(url)
                scraper._close_selenium()  # Cleanup
                return result
        raise ValueError(f"No scraper available for URL: {url}")


class ExcelExporter:
    """Export dữ liệu sản phẩm ra Excel"""

    def __init__(self):
        self.wb = Workbook()
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_align = Alignment(horizontal="center", vertical="center")

    def export(self, products: List[ProductData], output_path: str):
        """Export danh sách sản phẩm ra file Excel"""

        # Products sheet
        ws_products = self.wb.active
        ws_products.title = "Products"
        self._create_products_sheet(ws_products, products)

        # Variants sheet
        ws_variants = self.wb.create_sheet("Variants")
        self._create_variants_sheet(ws_variants, products)

        # Attributes sheet
        ws_attributes = self.wb.create_sheet("Attributes")
        self._create_attributes_sheet(ws_attributes, products)

        # Media sheet
        ws_media = self.wb.create_sheet("Media")
        self._create_media_sheet(ws_media, products)

        # Save
        self.wb.save(output_path)
        print(f"Exported to: {output_path}")

    def _set_header(self, ws, headers: List[str]):
        """Set header row với style"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_align

    def _create_products_sheet(self, ws, products: List[ProductData]):
        """Tạo sheet Products"""
        headers = [
            'name', 'sku_prefix', 'slug', 'brand_name', 'category_name',
            'base_price', 'short_description', 'description', 'is_featured',
            'status', 'meta_title', 'meta_description', 'tags'
        ]
        self._set_header(ws, headers)

        for row, p in enumerate(products, 2):
            ws.cell(row=row, column=1, value=p.name)
            ws.cell(row=row, column=2, value=p.sku_prefix)
            ws.cell(row=row, column=3, value=p.slug)
            ws.cell(row=row, column=4, value=p.brand_name)
            ws.cell(row=row, column=5, value=p.category_name)
            ws.cell(row=row, column=6, value=p.base_price)
            ws.cell(row=row, column=7, value=p.short_description)
            ws.cell(row=row, column=8, value=p.description[:32000] if p.description else "")  # Excel limit
            ws.cell(row=row, column=9, value=p.is_featured)
            ws.cell(row=row, column=10, value=p.status)
            ws.cell(row=row, column=11, value=p.meta_title or p.name)
            ws.cell(row=row, column=12, value=p.meta_description or p.short_description)
            ws.cell(row=row, column=13, value=p.tags)

        # Auto-fit columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    def _create_variants_sheet(self, ws, products: List[ProductData]):
        """Tạo sheet Variants"""
        headers = [
            'product_sku_prefix', 'sku', 'name', 'option_1_type', 'option_1_value',
            'option_2_type', 'option_2_value', 'option_2_color_code',
            'price', 'compare_at_price', 'stock_quantity', 'is_default', 'image_url'
        ]
        self._set_header(ws, headers)

        row = 2
        for p in products:
            for v in p.variants:
                ws.cell(row=row, column=1, value=p.sku_prefix)
                ws.cell(row=row, column=2, value=v.get('sku', ''))
                ws.cell(row=row, column=3, value=v.get('name', ''))
                ws.cell(row=row, column=4, value=v.get('option_1_type', ''))
                ws.cell(row=row, column=5, value=v.get('option_1_value', ''))
                ws.cell(row=row, column=6, value=v.get('option_2_type', ''))
                ws.cell(row=row, column=7, value=v.get('option_2_value', ''))
                ws.cell(row=row, column=8, value=v.get('option_2_color_code', ''))
                ws.cell(row=row, column=9, value=v.get('price', 0))
                ws.cell(row=row, column=10, value=v.get('compare_at_price', p.compare_at_price))
                ws.cell(row=row, column=11, value=v.get('stock_quantity', 0))
                ws.cell(row=row, column=12, value=v.get('is_default', False))
                ws.cell(row=row, column=13, value=v.get('image_url', ''))
                row += 1

    def _create_attributes_sheet(self, ws, products: List[ProductData]):
        """Tạo sheet Attributes"""
        headers = ['product_sku_prefix', 'attribute_name', 'value', 'display_group', 'display_order']
        self._set_header(ws, headers)

        row = 2
        for p in products:
            for attr in p.attributes:
                ws.cell(row=row, column=1, value=p.sku_prefix)
                ws.cell(row=row, column=2, value=attr.get('attribute_name', ''))
                ws.cell(row=row, column=3, value=attr.get('value', ''))
                ws.cell(row=row, column=4, value=attr.get('display_group', 'Thông tin chung'))
                ws.cell(row=row, column=5, value=attr.get('display_order', 0))
                row += 1

    def _create_media_sheet(self, ws, products: List[ProductData]):
        """Tạo sheet Media"""
        headers = ['product_sku_prefix', 'type', 'url', 'alt_text', 'display_order', 'is_primary']
        self._set_header(ws, headers)

        row = 2
        for p in products:
            for i, img_url in enumerate(p.images):
                ws.cell(row=row, column=1, value=p.sku_prefix)
                ws.cell(row=row, column=2, value='image')
                ws.cell(row=row, column=3, value=img_url)
                ws.cell(row=row, column=4, value=p.name)
                ws.cell(row=row, column=5, value=i + 1)
                ws.cell(row=row, column=6, value=i == 0)  # First image is primary
                row += 1


def main():
    parser = argparse.ArgumentParser(
        description='Cào dữ liệu sản phẩm từ web và lưu vào Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ví dụ:
  python product_scraper.py https://www.dienmayxanh.com/dien-thoai/iphone-15-pro-max
  python product_scraper.py https://cellphones.com.vn/iphone-15-pro-max.html --output iphone.xlsx
  python product_scraper.py url1 url2 url3 --output products.xlsx

Hỗ trợ các trang:
  - dienmayxanh.com
  - thegioididong.com
  - cellphones.com.vn
  - fptshop.com.vn
  - Và các trang khác (generic scraper)
        """
    )
    parser.add_argument('urls', nargs='+', help='URL sản phẩm cần cào')
    parser.add_argument('-o', '--output', default=None, help='File Excel output (default: products_YYMMDD_HHMMSS.xlsx)')
    parser.add_argument('-v', '--verbose', action='store_true', help='Hiển thị chi tiết')
    parser.add_argument('--selenium', '-s', action='store_true',
                        help='Dùng Selenium để cào trang có JavaScript (cần cài: pip install selenium)')

    args = parser.parse_args()

    # Check Selenium availability
    if args.selenium and not SELENIUM_AVAILABLE:
        print("⚠️  Selenium không khả dụng. Cài đặt: pip install selenium")
        print("   Và cần có ChromeDriver trong PATH")
        print("   Tiếp tục với requests...")
        args.selenium = False

    # Default output filename
    if not args.output:
        timestamp = datetime.now().strftime('%y%m%d_%H%M%S')
        args.output = f'products_{timestamp}.xlsx'

    # Scrape
    manager = ProductScraperManager(use_selenium=args.selenium)
    products = []

    for url in args.urls:
        print(f"\n{'='*60}")
        print(f"Scraping: {url}")
        print('='*60)

        try:
            product = manager.scrape(url)
            products.append(product)

            print(f"✓ Name: {product.name}")
            print(f"✓ Price: {product.base_price:,.0f}đ")
            print(f"✓ Brand: {product.brand_name}")
            print(f"✓ Category: {product.category_name}")
            print(f"✓ Images: {len(product.images)}")
            print(f"✓ Attributes: {len(product.attributes)}")
            print(f"✓ Variants: {len(product.variants)}")

            if args.verbose:
                print(f"\nAttributes:")
                for attr in product.attributes[:5]:
                    print(f"  - {attr['attribute_name']}: {attr['value']}")
                if len(product.attributes) > 5:
                    print(f"  ... và {len(product.attributes) - 5} thông số khác")

        except Exception as e:
            print(f"✗ Error: {e}")
            if args.verbose:
                import traceback
                traceback.print_exc()

    # Export
    if products:
        print(f"\n{'='*60}")
        print(f"Exporting {len(products)} products to Excel...")
        print('='*60)

        exporter = ExcelExporter()
        exporter.export(products, args.output)

        print(f"\n✓ Successfully exported to: {args.output}")
        print(f"✓ Products: {len(products)}")
        print(f"✓ Total variants: {sum(len(p.variants) for p in products)}")
        print(f"✓ Total attributes: {sum(len(p.attributes) for p in products)}")
        print(f"✓ Total images: {sum(len(p.images) for p in products)}")
    else:
        print("\n✗ No products scraped successfully")
        sys.exit(1)


if __name__ == '__main__':
    main()
